"""
Excel Import/Export Service - خدمة استيراد وتصدير Excel

Features:
- Export: customers, suppliers, products, invoices, reports
- Import: customers, suppliers, products from Excel
- Template generation for imports
- Multi-sheet support
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional


class ExcelService:
    """خدمة استيراد وتصدير Excel.

    يتطلب: pip install openpyxl
    """

    def __init__(self) -> None:
        self._openpyxl = None

    def _get_openpyxl(self):
        if self._openpyxl is None:
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                self._openpyxl = openpyxl
                self._Font = Font
                self._PatternFill = PatternFill
                self._Alignment = Alignment
                self._Border = Border
                self._Side = Side
            except ImportError:
                raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        return self._openpyxl

    def export_customers(self, customers: list[dict], filepath: str | Path) -> str:
        """تصدير العملاء إلى Excel."""
        wb = self._create_workbook("العملاء")
        ws = wb.active

        # Headers
        headers = ["الكود", "الاسم", "الهاتف", "البريد", "الرصيد", "حد الائتمان", "الحالة"]
        self._write_headers(ws, headers)

        # Data
        for i, c in enumerate(customers, 2):
            ws.cell(row=i, column=1, value=c.get("code", ""))
            ws.cell(row=i, column=2, value=c.get("name", ""))
            ws.cell(row=i, column=3, value=c.get("phone", ""))
            ws.cell(row=i, column=4, value=c.get("email", ""))
            ws.cell(row=i, column=5, value=float(c.get("current_balance", 0)))
            ws.cell(row=i, column=6, value=float(c.get("credit_limit", 0)))
            ws.cell(row=i, column=7, value="نشط" if c.get("is_active") else "معطّل")

        self._save_workbook(wb, filepath)
        return str(filepath)

    def export_products(self, products: list[dict], filepath: str | Path) -> str:
        """تصدير المنتجات إلى Excel."""
        wb = self._create_workbook("المنتجات")
        ws = wb.active

        headers = ["SKU", "الباركود", "الاسم", "التصنيف", "الوحدة", "سعر التكلفة", "سعر البيع", "الضريبة %", "الحد الأدنى"]
        self._write_headers(ws, headers)

        for i, p in enumerate(products, 2):
            ws.cell(row=i, column=1, value=p.get("sku", ""))
            ws.cell(row=i, column=2, value=p.get("barcode", ""))
            ws.cell(row=i, column=3, value=p.get("name", ""))
            ws.cell(row=i, column=4, value=p.get("category", ""))
            ws.cell(row=i, column=5, value=p.get("unit", ""))
            ws.cell(row=i, column=6, value=float(p.get("cost_price", 0)))
            ws.cell(row=i, column=7, value=float(p.get("sale_price", 0)))
            ws.cell(row=i, column=8, value=float(p.get("tax_rate", 15)))
            ws.cell(row=i, column=9, value=float(p.get("min_stock_level", 0)))

        self._save_workbook(wb, filepath)
        return str(filepath)

    def export_invoices(self, invoices: list[dict], filepath: str | Path) -> str:
        """تصدير الفواتير إلى Excel."""
        wb = self._create_workbook("الفواتير")
        ws = wb.active

        headers = ["رقم الفاتورة", "النوع", "الطرف", "التاريخ", "الإجمالي الفرعي", "الضريبة", "الإجمالي", "الحالة"]
        self._write_headers(ws, headers)

        for i, inv in enumerate(invoices, 2):
            ws.cell(row=i, column=1, value=inv.get("invoice_no", ""))
            ws.cell(row=i, column=2, value=inv.get("invoice_type", ""))
            ws.cell(row=i, column=3, value=inv.get("party_name", ""))
            ws.cell(row=i, column=4, value=str(inv.get("issue_date", "")))
            ws.cell(row=i, column=5, value=float(inv.get("subtotal", 0)))
            ws.cell(row=i, column=6, value=float(inv.get("tax_amount", 0)))
            ws.cell(row=i, column=7, value=float(inv.get("total", 0)))
            ws.cell(row=i, column=8, value=inv.get("status", ""))

        self._save_workbook(wb, filepath)
        return str(filepath)

    def export_trial_balance(self, lines: list[dict], filepath: str | Path) -> str:
        """تصدير ميزان المراجعة إلى Excel."""
        wb = self._create_workbook("ميزان المراجعة")
        ws = wb.active

        headers = ["الكود", "اسم الحساب", "النوع", "مدين", "دائن"]
        self._write_headers(ws, headers)

        for i, line in enumerate(lines, 2):
            ws.cell(row=i, column=1, value=line.get("account_code", ""))
            ws.cell(row=i, column=2, value=line.get("account_name", ""))
            ws.cell(row=i, column=3, value=line.get("account_type", ""))
            ws.cell(row=i, column=4, value=float(line.get("debit", 0)))
            ws.cell(row=i, column=5, value=float(line.get("credit", 0)))

        self._save_workbook(wb, filepath)
        return str(filepath)

    def import_customers(self, filepath: str | Path) -> list[dict]:
        """استيراد العملاء من Excel.

        Returns: list of customer dicts ready for CreateCustomerRequest.
        """
        openpyxl = self._get_openpyxl()
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb.active

        customers = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            customers.append({
                "code": str(row[0] or ""),
                "name": str(row[1] or ""),
                "phone": str(row[2] or ""),
                "email": str(row[3] or ""),
                "opening_balance": Decimal(str(row[4] or 0)),
                "credit_limit": Decimal(str(row[5] or 0)),
            })

        return customers

    def import_products(self, filepath: str | Path) -> list[dict]:
        """استيراد المنتجات من Excel."""
        openpyxl = self._get_openpyxl()
        wb = openpyxl.load_workbook(str(filepath))
        ws = wb.active

        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            products.append({
                "sku": str(row[0] or ""),
                "barcode": str(row[1] or ""),
                "name": str(row[2] or ""),
                "category": str(row[3] or ""),
                "unit": str(row[4] or "piece"),
                "cost_price": Decimal(str(row[5] or 0)),
                "sale_price": Decimal(str(row[6] or 0)),
                "tax_rate": Decimal(str(row[7] or 15)),
                "min_stock_level": Decimal(str(row[8] or 0)),
            })

        return products

    def generate_customer_template(self, filepath: str | Path) -> str:
        """توليد قالب Excel فارغ لاستيراد العملاء."""
        wb = self._create_workbook("قالب العملاء")
        ws = wb.active
        headers = ["الكود *", "الاسم *", "الهاتف", "البريد", "الرصيد الافتتاحي", "حد الائتمان"]
        self._write_headers(ws, headers)
        # Add example row
        ws.cell(row=2, column=1, value="C-001")
        ws.cell(row=2, column=2, value="شركة النور للتجارة")
        ws.cell(row=2, column=3, value="0501234567")
        ws.cell(row=2, column=4, value="info@alnoor.com")
        ws.cell(row=2, column=5, value=0)
        ws.cell(row=2, column=6, value=50000)
        self._save_workbook(wb, filepath)
        return str(filepath)

    def generate_product_template(self, filepath: str | Path) -> str:
        """توليد قالب Excel فارغ لاستيراد المنتجات."""
        wb = self._create_workbook("قالب المنتجات")
        ws = wb.active
        headers = ["SKU *", "الباركود", "الاسم *", "التصنيف", "الوحدة", "سعر التكلفة", "سعر البيع", "الضريبة %", "الحد الأدنى"]
        self._write_headers(ws, headers)
        ws.cell(row=2, column=1, value="PROD-001")
        ws.cell(row=2, column=2, value="1234567890")
        ws.cell(row=2, column=3, value="لابتوب Dell")
        ws.cell(row=2, column=4, value="إلكترونيات")
        ws.cell(row=2, column=5, value="piece")
        ws.cell(row=2, column=6, value=2000)
        ws.cell(row=2, column=7, value=2500)
        ws.cell(row=2, column=8, value=15)
        ws.cell(row=2, column=9, value=5)
        self._save_workbook(wb, filepath)
        return str(filepath)

    def _create_workbook(self, title: str):
        """إنشاء workbook جديد مع تنسيق."""
        openpyxl = self._get_openpyxl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        return wb

    def _write_headers(self, ws, headers: list[str]) -> None:
        """كتابة الترويسة مع تنسيق."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self._Font(bold=True, color="FFFFFF", size=12)
            cell.fill = self._PatternFill(start_color="6B8AFE", end_color="6B8AFE", fill_type="solid")
            cell.alignment = self._Alignment(horizontal="center", vertical="center")
            thin_border = self._Border(
                left=self._Side(style="thin"),
                right=self._Side(style="thin"),
                top=self._Side(style="thin"),
                bottom=self._Side(style="thin"),
            )
            cell.border = thin_border

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl_col_name(col)].width = 20

    def _save_workbook(self, wb, filepath: str | Path) -> None:
        """حفظ workbook."""
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(filepath))


def openpyxl_col_name(col: int) -> str:
    """تحويل رقم عمود إلى حرف (1 → A, 2 → B, ...)."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result
